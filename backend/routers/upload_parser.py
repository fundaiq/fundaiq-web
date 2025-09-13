# upload_parser.py

import pandas as pd
from openpyxl import load_workbook
import io
from collections import Counter

def format_column_headers(headers):
    formatted = []
    blank_counter = 1
    for h in headers:
        try:
            h_parsed = pd.to_datetime(h)
            formatted.append(h_parsed.strftime("%b-%Y"))
        except:
            if pd.notnull(h) and str(h).strip():
                formatted.append(str(h))
            else:
                formatted.append(f"Unnamed_{blank_counter}")
                blank_counter += 1
    counts = Counter()
    unique = []
    for h in formatted:
        counts[h] += 1
        unique.append(f"{h}_{counts[h]}" if counts[h] > 1 else h)
    return unique

def extract_table(df, start_label, start_row_offset, col_count=11):
    start_row = df[df.iloc[:, 0] == start_label].index[0]
    header_row = start_row + start_row_offset
    headers_raw = df.iloc[header_row, 1:col_count].tolist()
    formatted_headers = format_column_headers(headers_raw)

    # Keep only valid columns (non-Unnamed)
    valid_indices = [i for i, h in enumerate(formatted_headers) if not str(h).startswith("Unnamed_")]
    years = [formatted_headers[i] for i in valid_indices]
    data_rows = {}
    for i in range(header_row + 1, df.shape[0]):
        row_label = df.iloc[i, 0]
        row_values = [df.iloc[i, 1 + j] for j in valid_indices]
        if pd.isna(row_label):
            break
        data_rows[str(row_label).strip()] = row_values[:len(years)]
    return data_rows, years

def extract_meta(df):
    meta_start = df[df.iloc[:, 0] == "META"].index[0]
    df_meta = df.iloc[meta_start+1:, 0:2].dropna()
    df_meta.columns = ["Label", "Value"]
    raw = df_meta.set_index("Label")["Value"].to_dict()
    safe_meta = {
        k: str(v) if isinstance(v, (pd.Timestamp, pd._libs.tslibs.nattype.NaTType)) else v
        for k, v in raw.items()
    }
    return safe_meta

def extract_quarters(df):
    try:
        df_quarters, _ = extract_table(df, "QUARTERS", 1)
        return df_quarters
    except:
        return {}

def normalize_table(table):
    return {
        k: [v if v is not None else 0 for v in row]
        for k, row in table.items()
    }

def make_json_safe(obj):
    if isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    elif isinstance(obj, (pd.Timestamp, pd._libs.tslibs.nattype.NaTType)):
        return str(obj)
    else:
        return obj

def parse_excel(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Data Sheet"]
    df_all = pd.DataFrame(ws.values)

    company_name = ws["B1"].value or "Unknown Company"
    meta = extract_meta(df_all)
    #print(f"ℹ️ [BACKEND DEBUG] Meta data: {meta}")

    pnl, pnl_years = extract_table(df_all, "PROFIT & LOSS", 1)
    ##print(f"ℹ️ [BACKEND DEBUG] Meta data: {pnl_years}")
    ##print(f"ℹ️ [BACKEND DEBUG] Meta data: {pnl}")

    bs, bs_years = extract_table(df_all, "BALANCE SHEET", 1)
    ##print(f"ℹ️ [BACKEND DEBUG] Meta data: {bs_years}")
    ##print(f"ℹ️ [BACKEND DEBUG] Meta data: {bs}")

    cf, cf_years = extract_table(df_all, "CASH FLOW:", 1)
    ##print(f"ℹ️ [BACKEND DEBUG] Meta data: {cf_years}")
    ##print(f"ℹ️ [BACKEND DEBUG] Meta data: {cf}")

    quarters, quarters_years = extract_table(df_all, "Quarters", 1)
    
    

    pnl = normalize_table(pnl)
    bs = normalize_table(bs)
    cf = normalize_table(cf)
    quarters = normalize_table(quarters)
    #print(f"ℹ️ [BACKEND DEBUG] Parser PnL : {pnl}")
    #print(f"ℹ️ [BACKEND DEBUG] Parser Quarters : {quarters}")

    return make_json_safe({
        "company_name": company_name,
        "meta": meta,
        "pnl": pnl,
        "balance_sheet": bs,
        "cashflow": cf,
        "quarters": quarters,
        "years": [y for y in pnl_years if y in bs_years and y in cf_years],
        "qtrs": [y for y in quarters_years]
    })